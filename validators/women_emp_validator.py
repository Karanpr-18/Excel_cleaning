
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import os
import logging

class WomenEmpValidator:
    """Women Empowerment Excel file validator"""
    
    def __init__(self):
        # Validation rules for Women Empowerment data
        self.validation_rules = {
            "State": ["not null"],
            "District": ["not null"],
            "Block": ["not null"],
            "Village": ["not null"],
            "Project": ["not null"],
            "User Name(FE)": ["not null", "name_only_alphabets"],
            "Cast": ["not null"],
            "Economic Status": ["not null"],
            "Marital Status": ["not null"],
            "Registration Date": ["not null"],
            "Education": ["not null"],
            "Women Name": ["not null", "name_only_alphabets"],
            "Husband / Father Name": ["not null", "name_only_alphabets"],
            "Mother Name": ["not null", "name_only_alphabets"],
            "Phone No.": ["phone_validation"],
            "Any ID Proof Details": ["not null"],
            "ID Proof No.": ["conditional_id_proof"],
            "Ration Card": ["not null"],
            "Ration Card linked PDS": ["not null"],
            "Bank Account No.": ["not null"],
            "Monthly Individual Income": ["not null"],
            "Monthly Household Income": ["not null"],
            "Is Life Skills Training": ["not null"],
            "Start Business": ["not null"],
            "Business": ["not null"],
            "Business When": ["not null"],
            "Status Business": ["not null"],
            "Village Population": ["not null", "numeric"],
            "Business Idea": ["not null"],
            "Business Type": ["not null"],
            "Procure Business": ["not null"],
            "Current Business": ["not null"],
            "Regular Financial Business": ["not null"],
            "How Regular Financial": ["not null"],
            "Setting Business Type": ["not null"],
            "Potential Customers": ["not null"],
            "Business Distance": ["not null"],
            "How Far Bussiness": ["not null"],
            "Planning Business": ["not null"],
            "Support Business": ["not null"],
            "Support Type": ["not null"],
            "Not Provided Support": ["not null"],
            "Own Smart Phone": ["not null"],
            "Use Smart Phone": ["not null"],
            "Supply Chain": ["not null"],
            "Date Of Business Inauguration": ["not null"],
            "Aadhaar Card Details": ["not null"],
            "Aadhaar No.": ["conditional_aadhaar"]
        }
        
        # Fill color for errors
        self.red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    def validate_excel(self, file_path, unique_id, download_folder='downloads'):
        """
        Validate Excel file using Women Empowerment method and generate output files
        Returns dict with paths to generated files or None if error
        """
        try:
            # Read the Excel file - use the first sheet available
            excel_file = pd.ExcelFile(file_path)
            first_sheet_name = excel_file.sheet_names[0]
            data_df = pd.read_excel(file_path, sheet_name=first_sheet_name)
            
            # Load workbook for styling
            wb = load_workbook(file_path)
            ws = wb[first_sheet_name]
            
            validation_errors = []
            
            # Process each column according to validation rules
            for col_name, rules in self.validation_rules.items():
                if col_name not in data_df.columns:
                    continue
                
                # Find column letter in worksheet
                col_letter = None
                for cell in ws[1]:
                    if cell.value == col_name:
                        col_letter = cell.column_letter
                        break
                if col_letter is None:
                    continue
                
                # Validate each cell in the column
                for i, value in enumerate(data_df[col_name], start=2):
                    has_error = False
                    error_reason = ""
                    
                    # Apply validation rules
                    for rule in rules:
                        rule_lower = rule.lower()
                        
                        if rule_lower == "not null":
                            if pd.isnull(value) or (isinstance(value, str) and str(value).strip() == ""):
                                has_error = True
                                error_reason = "Mandatory field is empty"
                        elif rule_lower == "name_only_alphabets":
                            if value and not str(value).replace(" ", "").isalpha():
                                has_error = True
                                error_reason = "Name field should contain only alphabets"
                        elif rule_lower == "phone_validation":
                            if value and not (str(value).isdigit() and len(str(value)) == 10):
                                has_error = True
                                error_reason = "Phone No. should be exactly 10 digits"
                        elif rule_lower == "numeric":
                            if pd.isnull(value) or str(value).strip() == "" or not str(value).isdigit():
                                has_error = True
                                error_reason = "Village Population should be numeric and not empty"
                        elif rule_lower == "conditional_id_proof":
                            # Check if ID Proof Details is provided
                            if "Any ID Proof Details" in data_df.columns:
                                id_proof_details = data_df["Any ID Proof Details"].iloc[i - 2]
                                if pd.notna(id_proof_details) and str(id_proof_details).strip() != "":
                                    if pd.isnull(value) or str(value).strip() == "":
                                        has_error = True
                                        error_reason = "ID Proof No. required when Any ID Proof Details is provided"
                        elif rule_lower == "conditional_aadhaar":
                            # Check Aadhaar validation based on Aadhaar Card Details
                            if "Aadhaar Card Details" in data_df.columns:
                                aadhaar_details = data_df["Aadhaar Card Details"].iloc[i - 2]
                                if pd.notna(aadhaar_details) and str(aadhaar_details).strip().lower() == "yes":
                                    if pd.isnull(value) or str(value).strip() == "":
                                        has_error = True
                                        error_reason = "Aadhaar No. required when Aadhaar Card Details is yes"
                                    elif not (str(value).isdigit() and len(str(value)) == 12):
                                        has_error = True
                                        error_reason = "Aadhaar No. should be exactly 12 digits when Aadhaar Card Details is yes"
                        
                        if has_error:
                            break
                    
                    # Mark errors in worksheet and record them
                    if has_error:
                        cell_ref = f"{col_letter}{i}"
                        ws[cell_ref].fill = self.red_fill
                        validation_errors.append({
                            "Row": i,
                            "Column": col_name,
                            "Cell": cell_ref,
                            "Value": value,
                            "Error": error_reason,
                        })
            
            # Generate output file paths
            output_path = os.path.join(download_folder, f"{unique_id}_Validated_Output_WomenEmp.xlsx")
            report_path = os.path.join(download_folder, f"{unique_id}_Validation_Report_WomenEmp.xlsx")
            
            # Save highlighted workbook
            wb.save(output_path)
            wb.close()  # Explicitly close the workbook to release file handles
            
            # Create validation report
            if validation_errors:
                report_df = pd.DataFrame(validation_errors)
            else:
                # Create empty DataFrame with proper column structure
                report_df = pd.DataFrame({
                    "Row": pd.Series([], dtype='int64'),
                    "Column": pd.Series([], dtype='object'),
                    "Cell": pd.Series([], dtype='object'),
                    "Value": pd.Series([], dtype='object'),
                    "Error": pd.Series([], dtype='object')
                })
            
            report_df.to_excel(report_path, index=False)
            
            logging.info(f"Women Empowerment validation complete. Generated {len(validation_errors)} error records.")
            
            return {
                'validated_output': output_path,
                'validation_report': report_path
            }
            
        except Exception as e:
            logging.error(f"Error during Women Empowerment Excel validation: {str(e)}")
            return None
