import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import os
import logging
from functools import lru_cache
from datetime import datetime, timedelta


class KadamPlusValidator:
    def __init__(self):
        self.validation_rules = {
            "Student's First Name": ["not null", "no_special_chars"],
            "Student's Age": ["not null", "numeric", "age_range"],  # Changed to age_range
            "Student's Date of Birth": ["not null", "date"],
            "Father's Name": ["not null", "no_special_chars"],
            "Father's Age": ["not null", "numeric"],
            "Mother's Name": ["not null"],
            "Mother's Age": ["not null", "numeric"],
            "Contact No.": ["not null", "numeric"],
            "House Address": ["not null"],
            "Pincode": ["not null", "numeric"],
            "People living in house": ["not null", "numeric"],
            "Cast": ["not null", "no_special_chars"],
            "Religion": ["not null", "no_special_chars"],
            "Parents' Monthly Income":["not null"],
            "Parents' Monthly Expenditure": ["not null"],
            "Trio No.":["not_zero"],
            "Baseline Math": ["not null", "numeric"],
            "Baseline English": ["not null", "numeric"],
            "Baseline EVS": ["not null", "numeric"],
            "Baseline Hindi": ["not null", "numeric"],
            "Baseline Total": ["numeric"],
            "Endline Math": ["numeric"],
            "Endline English": ["numeric"],
            "Endline EVS": ["numeric"],
            "Endline Hindi": ["numeric"],
            "Endline Total": ["numeric"],
            "Grade Test 1": ["numeric"],
            "Grade Test 2": ["numeric"],
            "Grade Test 3": ["numeric"],
            "Grade Test 4": ["numeric"],
            "Grade Test 5": ["numeric"],
            "Enrolment Grade": ["not null"],
            "No. of Steps Completed": ["numeric", "not_zero"],
            "Date of Admission": ["date"]
        }
        self.red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # Age limits
        self.min_age = 6.5  # 6 years 6 months
        self.max_age = 14 + (1/12)  # 14 years 1 month (approximately 14.083 years)
        
        # Pre-compile regex patterns for better performance
        self.name_pattern = re.compile(r"^[A-Za-z ]*$")
        self.phone_decimal_pattern = re.compile(r"\d+\.0")
        self.non_digit_pattern = re.compile(r"\D")
        
        # Cache for frequently used data
        self.grade_test_columns = frozenset([
            "Grade Test 1", "Grade Test 2", "Grade Test 3", "Grade Test 4", "Grade Test 5"
        ])
        self.baseline_subjects = frozenset([
            "Baseline Math", "Baseline English", "Baseline EVS", "Baseline Hindi"
        ])
        self.endline_subjects = frozenset([
            "Endline Math", "Endline English", "Endline EVS", "Endline Hindi"
        ])
        self.all_subjects = self.baseline_subjects | self.endline_subjects
        self.total_columns = frozenset(["Baseline Total", "Endline Total"])
        self.parent_age_columns = frozenset(["Father's Age", "Mother's Age"])
        
        # Grade-based limits (immutable)
        self.grade_max = {"1": 10, "2": 20, "3": 30, "4": 40}
        self.grade_max_total = {"1": 40, "2": 80, "3": 120, "4": 160}

    @lru_cache(maxsize=1000)
    def _is_valid_number(self, value_str):
        """Cached number validation"""
        try:
            float(value_str)
            return True
        except (ValueError, TypeError):
            return False

    @lru_cache(maxsize=500)
    def _is_valid_date(self, value_str):
        """Cached date validation"""
        try:
            return pd.to_datetime(value_str, errors='coerce') is not pd.NaT
        except:
            return False

    def _is_null_or_empty(self, value):
        """Fast null/empty check"""
        return pd.isnull(value) or (isinstance(value, str) and not value.strip())

    def _validate_student_age(self, age_value, row):
        """Validate student age (must be between 6y6m and 14y1m)"""
        try:
            age = float(age_value)
            
            # Check if Date of Birth and Date of Admission are available for more accurate calculation
            if 'Student\'s Date of Birth' in row and 'Date of Admission' in row:
                dob = row['Student\'s Date of Birth']
                admission_date = row['Date of Admission']
                
                if pd.notnull(dob) and pd.notnull(admission_date):
                    try:
                        dob_parsed = pd.to_datetime(dob)
                        admission_parsed = pd.to_datetime(admission_date)
                        
                        # Calculate age at admission in years
                        age_at_admission_days = (admission_parsed - dob_parsed).days
                        age_at_admission_years = age_at_admission_days / 365.25
                        
                        if age_at_admission_years < self.min_age:
                            return True, f"Student age at admission ({age_at_admission_years:.1f} years) must be at least 6 years 6 months"
                        
                        if age_at_admission_years > self.max_age:
                            return True, f"Student age at admission ({age_at_admission_years:.1f} years) must not exceed 14 years 1 month"
                            
                        return False, ""  # Valid age range
                        
                    except:
                        # Fall back to simple age check if date parsing fails
                        pass
            
            # Simple age check if dates are not available or parsing failed
            if age < self.min_age:
                return True, f"Student age ({age} years) must be at least 6 years 6 months"
            
            if age > self.max_age:
                return True, f"Student age ({age} years) must not exceed 14 years 1 month"
                
        except (ValueError, TypeError):
            return True, "Invalid student age"
        
        return False, ""

    def _validate_contact_number(self, value):
        """Optimized contact number validation"""
        if pd.isnull(value):
            return False, ""
        
        try:
            value_str = str(value).strip()
            # Handle decimal format
            if self.phone_decimal_pattern.fullmatch(value_str):
                value_str = value_str.split('.')[0]
            # Remove non-digits
            value_str = self.non_digit_pattern.sub("", value_str)
            
            if len(value_str) != 10:
                return True, "Contact No. must be exactly 10 digits"
                
        except Exception:
            return True, "Invalid contact number"
        
        return False, ""

    def _validate_parent_age(self, value, column_name):
        """Validate parent age with caching"""
        try:
            age = float(value)
            if age < 20:
                return True, f"{column_name} should not be less than 20"
        except (ValueError, TypeError):
            return True, f"Invalid age in {column_name}"
        return False, ""

    def _validate_grade_test(self, value):
        """Validate grade test scores"""
        if pd.isnull(value):
            return False, ""
        
        try:
            score = float(value)
            if score < 40:
                return True, "Grade test marks cannot be less than 40"
        except (ValueError, TypeError):
            return True, "Invalid grade test score"
        return False, ""

    def _validate_subject_score(self, value, column_name, grade):
        """Validate subject scores with grade limits"""
        try:
            if column_name in self.baseline_subjects and pd.isnull(value):
                return True, f"{column_name} cannot be null"
                
            if pd.notnull(value) and grade:
                subject_score = float(value)
                max_score = self.grade_max.get(str(grade).strip())
                if max_score and subject_score > max_score:
                    return True, f"{column_name} exceeds max allowed ({max_score}) for Grade {grade}"
                    
        except (ValueError, TypeError):
            return True, "Invalid subject score"
        return False, ""

    def _validate_total_score(self, value, column_name, grade, baseline_total=None):
        """Validate total scores with grade limits and baseline comparison"""
        try:
            # Allow null values for totals
            if pd.isnull(value):
                return False, ""
                
            if grade:
                total_score = float(value)
                max_total = self.grade_max_total.get(str(grade).strip())
                if max_total and total_score > max_total:
                    return True, f"{column_name} exceeds max allowed total ({max_total}) for Grade {grade}"
            
            # Check endline vs baseline comparison
            if column_name == "Endline Total" and baseline_total is not None and pd.notnull(baseline_total):
                endline_total = float(value)
                baseline_val = float(baseline_total)
                if endline_total < baseline_val:
                    return True, f"Endline Total ({endline_total}) cannot be lower than Baseline Total ({baseline_val})"
                    
        except (ValueError, TypeError):
            return True, "Invalid total score"
        return False, ""

    def validate_cell(self, value, rules, column_name=None, row=None):
        """Optimized cell validation with early returns"""
        
        # Fast path for basic rule validation
        for rule in rules:
            if rule == "not null":
                if self._is_null_or_empty(value):
                    return True, "Value is null or empty"
            elif rule == "numeric":
                if not self._is_valid_number(str(value)):
                    return True, "Value is not numeric"
            elif rule == "date":
                if not self._is_valid_date(str(value)):
                    return True, "Invalid date"
            elif rule == "no_special_chars":
                if not self.name_pattern.match(str(value)):
                    return True, "Special characters found"
            elif rule == "not_zero":
                try:
                    if float(value) == 0:
                        return True, "Value cannot be zero"
                except (ValueError, TypeError):
                    return True, "Invalid value for zero check"
            elif rule == "age_range":
                if column_name == "Student's Age":
                    return self._validate_student_age(value, row)

        # Optimized column-specific validations
        if column_name == "Contact No.":
            return self._validate_contact_number(value)
            
        elif column_name in self.parent_age_columns:
            return self._validate_parent_age(value, column_name)
            
        elif column_name == "People living in house":
            try:
                if float(value) <= 1:
                    return True, "People living in house must be more than 1"
            except (ValueError, TypeError):
                return True, "Invalid number for People living in house"
                
        elif column_name in self.grade_test_columns:
            return self._validate_grade_test(value)
            
        elif column_name in self.all_subjects:
            grade = row.get('Enrolment Grade') if row else None
            return self._validate_subject_score(value, column_name, grade)
            
        elif column_name in self.total_columns:
            grade = row.get('Enrolment Grade') if row else None
            baseline_total = row.get('Baseline Total') if row else None
            return self._validate_total_score(value, column_name, grade, baseline_total)

        return False, ""

    def validate_excel(self, file_path, unique_id, download_folder='downloads'):
        try:
            # Read Excel file once
            df = pd.read_excel(file_path)
            wb = load_workbook(file_path)
            ws = wb.active

            # Pre-calculate column mappings
            col_letter_map = {}
            for cell in ws[1]:
                if cell.value in self.validation_rules:
                    col_letter_map[cell.value] = cell.column_letter

            errors = []
            
            # Batch process rows for better performance
            for idx, row in df.iterrows():
                excel_row = idx + 2
                row_dict = row.to_dict()  # Convert once per row
                
                for col, rules in self.validation_rules.items():
                    if col not in df.columns or col not in col_letter_map:
                        continue
                        
                    value = row_dict[col]
                    has_error, reason = self.validate_cell(value, rules, column_name=col, row=row_dict)
                    
                    if has_error:
                        cell_ref = f"{col_letter_map[col]}{excel_row}"
                        ws[cell_ref].fill = self.red_fill
                        errors.append({
                            "Row": excel_row, 
                            "Column": col, 
                            "Cell": cell_ref, 
                            "Value": value, 
                            "Error": reason
                        })

            # Create output directory
            os.makedirs(download_folder, exist_ok=True)
            validated_path = os.path.join(download_folder, f"{unique_id}_Validated_Output_KadamPlus.xlsx")
            report_path = os.path.join(download_folder, f"{unique_id}_Validation_Report_KadamPlus.xlsx")

            # Save files
            wb.save(validated_path)
            wb.close()

            # Create report DataFrame efficiently
            if errors:
                error_df = pd.DataFrame(errors)
            else:
                error_df = pd.DataFrame(columns=["Row", "Column", "Cell", "Value", "Error"])
            
            error_df.to_excel(report_path, index=False)

            logging.info(f"Validation complete. {len(errors)} error(s) found.")
            return {'validated_output': validated_path, 'validation_report': report_path}

        except Exception as e:
            logging.error(f"Validation failed: {str(e)}")
            return None
