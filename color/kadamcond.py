import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

file_path = "Kadam Haryana compile report  2024-2025 02052025.xlsx"
data_df = pd.read_excel(file_path, sheet_name='Compile Report')

validation_rules = {
    "Student's First Name": ["not null", "no_special_chars"],
    "Student's Age": ["not null", "numeric"],
    "Student's Date of Birth": ["not null", "date"],
    "Father's Age": ["not null", "numeric"],
    "Father's Occupation": ["not null", "no_special_chars"],
    "Father's Education": ["not null"],
    "Mother's Name": ["not null"],
    "Mother's Age": ["not null", "numeric"],
    "Mother's Occupation": ["not null", "no_special_chars"],
    "How long are you planning to stay in this area?": ["not null"],
    "Contact No.": ["not null", "numeric"],
    "House Address": ["not null"],
    "Pincode": ["not null", "numeric"],
    "People living in house": ["not null", "numeric"],
    "Cast": ["not null", "no_special_chars"],
    "Religion": ["not null", "no_special_chars"],
    "Parents' Monthly Income": ["not null"],
    "Parents' Monthly Expenditure": ["not null"],
    "Baseline Math": ["not null", "numeric"],
    "Baseline English": ["not null", "numeric"],
    "Baseline EVS": ["not null", "numeric"],
    "Baseline Hindi": ["not null", "numeric"],
    "Baseline Total": ["not null", "numeric"],
    "Baseline Percentage": ["not null", "numeric"],
    "Grade Test 1": ["not null", "numeric"],
    "Grade Test 2": ["not null", "numeric"],
    "Grade Test 3": ["not null", "numeric"],
    "Grade Test 4": ["not null", "numeric"],
    "Grade Test 5": ["not null", "numeric"],
    "Endline Math": ["not null", "numeric"],
    "Endline English": ["not null", "numeric"],
    "Endline EVS": ["not null", "numeric"],
    "Endline Hindi": ["not null", "numeric"],
    "Endline Total": ["not null", "numeric"],
    "Mainstream Institution Name": ["not null"],
    "Mainstream Institution Address": ["not null"],
    "School DISE Code": ["not null", "numeric"],
    "Mainstream Grade": ["not null"],
    "Mainstream Section": ["not null"],
    "Child SR given by the Institution": ["not null"],
    "State": ["not null"],
    "District": ["not null"],
    "Mainstream Date": ["not null", "date"],
    "Current Grade After Mainstream": ["not null"],
}

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

wb = load_workbook(file_path)
ws = wb['Compile Report']

validation_errors = []

def get_max_marks(age, total=False):
    try:
        age_num = int(age)
    except (ValueError, TypeError):
        return None
    if total:
        if age_num == 6:
            return 40
        elif age_num == 7:
            return 80
        elif age_num == 8:
            return 120
        elif 9 <= age_num <= 14:
            return 160
        else:
            return None
    else:
        if age_num == 6:
            return 10
        elif age_num == 7:
            return 20
        elif age_num == 8:
            return 30
        elif 9 <= age_num <= 14:
            return 40
        else:
            return None

original_limit_columns = [
    "Baseline English",
    "Baseline EVS",
    "Baseline Hindi",
    "Baseline Math",
    "Endline Math",
    "Endline English",
    "Endline EVS",
    "Endline Hindi",
]

new_limit_columns = ["Baseline Total", "Endline Total"]

grade_test_columns = [
    "Grade Test 1",
    "Grade Test 2",
    "Grade Test 3",
    "Grade Test 4",
    "Grade Test 5",
]

for col_name, rules in validation_rules.items():
    if col_name not in data_df.columns:
        continue

    col_letter = None
    for cell in ws[1]:
        if cell.value == col_name:
            col_letter = cell.column_letter
            break
    if col_letter is None:
        continue

    for i, value in enumerate(data_df[col_name], start=2):
        has_error = False
        error_reason = ""

        for rule in rules:
            rule_lower = rule.lower()

            if rule_lower == "not null":
                if pd.isnull(value) or (isinstance(value, str) and str(value).strip() == ""):
                    has_error = True
                    error_reason = "Value is null or empty"
            elif rule_lower == "numeric":
                try:
                    float(value)
                except Exception:
                    has_error = True
                    error_reason = "Value is not numeric"
            elif rule_lower == "date":
                if pd.to_datetime(value, errors='coerce') is pd.NaT:
                    has_error = True
                    error_reason = "Value is not a valid date"
            elif rule_lower == "no_special_chars":
                if not re.match(r"^[A-Za-z ]*$", str(value)):
                    has_error = True
                    error_reason = "Value contains special characters"

            if has_error:
                break

        # Color age less than 6 or greater than 14 as error
        if not has_error and col_name == "Student's Age":
            try:
                age_val = float(value)
                if age_val < 6 or age_val > 14:
                    has_error = True
                    error_reason = "Age is less than 6 or greater than 14"
            except Exception:
                # numeric error already caught
                pass

        if not has_error:
            age_value = data_df["Student's Age"].iloc[i - 2]

            if col_name in original_limit_columns:
                max_marks = get_max_marks(age_value, total=False)
            elif col_name in new_limit_columns:
                max_marks = get_max_marks(age_value, total=True)
            elif col_name in grade_test_columns:
                max_marks = 40  # fixed max for all grade tests regardless of age
            else:
                max_marks = None

            if max_marks is not None:
                try:
                    val = float(value)
                    if val > max_marks:
                        has_error = True
                        error_reason = f"Value exceeds max allowed marks ({max_marks})"
                except Exception:
                    # numeric error already caught
                    pass

        if has_error:
            cell_ref = f"{col_letter}{i}"
            ws[cell_ref].fill = red_fill
            validation_errors.append({
                "Row": i,
                "Column": col_name,
                "Cell": cell_ref,
                "Value": value,
                "Error": error_reason,
            })
            continue

wb.save("Validated_Output.xlsx")

report_df = pd.DataFrame(validation_errors)
if report_df.empty:
    report_df = pd.DataFrame(columns=["Row", "Column", "Cell", "Value", "Error"])

report_df.to_excel("Validation_Report.xlsx", index=False)

print("Validation complete.")
print("Highlighted data saved as 'Validated_Output.xlsx'.")
print("Validation report saved as 'Validation_Report.xlsx'.")