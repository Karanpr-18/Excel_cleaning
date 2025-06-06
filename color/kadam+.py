import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

file_path = "Kadam+test file.xlsx"
data_df = pd.read_excel(file_path, sheet_name='Consolidated')

validation_rules = {
    "Student's First Name": ["not null", "no_special_chars"],
    "Student's Date of Birth": ["not null", "date"],
    "Father's Age": ["not null", "numeric"],
    "Father's Occupation": ["not null"],
    "Father's Education": ["not null"],
    "Mother's Name": ["not null"],
    "Mother's Age": ["not null", "numeric"],
    "Mother's Occupation": ["not null"],
    "How long are you planning to stay in this area?": ["not null"],
    "Contact No.": ["not null", "numeric"],
    "House Address": ["not null"],
    "Pincode": ["not null", "numeric"],
    "People living in house": ["not null", "numeric"],
    "Cast": ["not null", "no_special_chars"],
    "Religion": ["not null", "no_special_chars"],
    "Parents' Monthly Income": ["not null"],
    "Parents' Monthly Expenditure": ["not null"],
    "Baseline Math": ["not null", "numeric"],
    "Baseline English": ["not null", "numeric"],
    "Baseline EVS": ["not null", "numeric"],
    "Baseline Hindi": ["not null", "numeric"],
    "Baseline Total": ["not null", "numeric"],
    "Baseline Percentage": ["not null", "numeric"],
    "Grade Test 1": ["not null", "numeric"],
    "Grade Test 2": ["not null", "numeric"],
    "Grade Test 3": ["not null", "numeric"],
    "Grade Test 4": ["not null", "numeric"],
    "Grade Test 5": ["not null", "numeric"],
    "Endline Math": ["not null", "numeric"],
    "Endline English": ["not null", "numeric"],
    "Endline EVS": ["not null", "numeric"],
    "Endline Hindi": ["not null", "numeric"],
    "Endline Total": ["not null", "numeric"],
}

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

wb = load_workbook(file_path)
ws = wb['Consolidated']

validation_errors = []

baseline_subjects = [
    "Baseline English",
    "Baseline EVS",
    "Baseline Hindi",
    "Baseline Math",
]

endline_subjects = [
    "Endline Math",
    "Endline English",
    "Endline EVS",
    "Endline Hindi",
]

new_limit_columns = ["Baseline Total", "Endline Total"]

grade_test_columns = [
    "Grade Test 1",
    "Grade Test 2",
    "Grade Test 3",
    "Grade Test 4",
    "Grade Test 5",
]

valid_grades = {"1st", "2nd", "3rd", "4th", "5th"}
grade_order = {"1st": 1, "2nd": 2, "3rd": 3, "4th": 4, "5th": 5}

max_marks_by_grade = {
    1: 10,
    2: 20,
    3: 30,
    4: 40,
}

total_max_marks_by_grade = {
    1: 40,
    2: 80,
    3: 120,
    4: 160,
    5: 160,  # Assuming max allowed mark as 160 for grade 5 and above
}

for col_name, rules in validation_rules.items():
    if col_name not in data_df.columns:
        continue

    col_letter = None
    for cell in ws[1]:
        if cell.value == col_name:
            col_letter = cell.column_letter
            break
    if col_letter is None:
        continue

    for i, value in enumerate(data_df[col_name], start=2):
        has_error = False
        error_reason = ""

        for rule in rules:
            rule_lower = rule.lower()

            if rule_lower == "not null":
                if pd.isnull(value) or (isinstance(value, str) and str(value).strip() == ""):
                    has_error = True
                    error_reason = "Value is null or empty"
                    break
            elif rule_lower == "numeric":
                try:
                    float(value)
                except (ValueError, TypeError):
                    has_error = True
                    error_reason = "Value is not numeric"
                    break
            elif rule_lower == "date":
                if pd.to_datetime(value, errors='coerce') is pd.NaT:
                    has_error = True
                    error_reason = "Value is not a valid date"
                    break
            elif rule_lower == "no_special_chars":
                if not re.match(r"^[A-Za-z ]*$", str(value)):
                    has_error = True
                    error_reason = "Value contains special characters"
                    break

        if not has_error:
            grade_value = None
            if "Current Grade After Mainstream" in data_df.columns:
                grade_val_raw = data_df["Current Grade After Mainstream"].iloc[i - 2]
                if pd.isnull(grade_val_raw):
                    grade_value = 5  # assume no restriction if grade not provided
                else:
                    grade_val_str = str(grade_val_raw).strip().lower()
                    if grade_val_str.endswith(("st", "nd", "rd", "th")):
                        grade_value = grade_order.get(grade_val_str, None)
                    elif grade_val_str.isdigit():
                        grade_num = int(grade_val_str)
                        if 1 <= grade_num <= 5:
                            grade_value = grade_num
                    else:
                        grade_value = 5
            else:
                grade_value = 5  # fallback no restriction

            max_marks = None

            # Important: For Baseline Total and Endline Total - use strict grade-based limits from total_max_marks_by_grade
            if col_name in new_limit_columns:
                max_marks = total_max_marks_by_grade.get(grade_value, 40)
            elif col_name in baseline_subjects or col_name in endline_subjects:
                max_marks = max_marks_by_grade.get(grade_value, None)
            elif col_name in grade_test_columns:
                max_marks = 40

            if max_marks is not None:
                try:
                    val = float(value)
                    if val > max_marks:
                        has_error = True
                        error_reason = f"Value exceeds max allowed marks ({max_marks})"
                except (ValueError, TypeError):
                    pass

        if has_error:
            cell_ref = f"{col_letter}{i}"
            ws[cell_ref].fill = red_fill
            validation_errors.append({
                "Row": i,
                "Column": col_name,
                "Cell": cell_ref,
                "Value": value,
                "Error": error_reason,
            })

wb.save("Validated_Output_kadam+.xlsx")

report_df = pd.DataFrame(validation_errors)
if report_df.empty:
    report_df = pd.DataFrame(columns=["Row", "Column", "Cell", "Value", "Error"])

report_df.to_excel("Validation_Report_kadam+.xlsx", index=False)

print("Validation complete.")
print("Highlighted data saved as 'Validated_Output_kadam+.xlsx'.")
print("Validation report saved as 'Validation_Report_kadam+.xlsx'.")